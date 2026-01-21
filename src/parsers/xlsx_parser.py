from openpyxl import load_workbook
from .base_parser import BaseParser


class XLSXParser(BaseParser):
    """Parser for XLSX documents"""

    def parse(self) -> str:
        """Extract data from Excel and convert to markdown"""
        markdown_lines = []
        markdown_lines.append(f"# {self.file_path.stem}\n")
        markdown_lines.append(f"*Source: {self.file_path.name}*\n")
        markdown_lines.append("---\n")

        try:
            workbook = load_workbook(self.file_path, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Add sheet header
                markdown_lines.append(f"\n## Sheet: {sheet_name}\n")

                # Get all data from sheet
                data = []
                for row in sheet.iter_rows(values_only=True):
                    # Skip completely empty rows
                    if any(cell is not None for cell in row):
                        data.append(row)

                if not data:
                    markdown_lines.append("*Empty sheet*\n")
                    continue

                # Convert to markdown table
                markdown_lines.append(self._data_to_markdown(data))
                markdown_lines.append("\n")

        except Exception as e:
            markdown_lines.append(f"\n**Error parsing XLSX:** {str(e)}\n")

        return "\n".join(markdown_lines)

    def _data_to_markdown(self, data) -> str:
        """Convert spreadsheet data to markdown table"""
        if not data:
            return ""

        markdown_table = []

        # Determine max column count
        max_cols = max(len(row) for row in data)

        # Normalize all rows to have the same number of columns
        normalized_data = []
        for row in data:
            normalized_row = list(row) + [None] * (max_cols - len(row))
            normalized_data.append(normalized_row)

        # Header row
        header = [str(cell) if cell is not None else "" for cell in normalized_data[0]]
        markdown_table.append("| " + " | ".join(header) + " |")
        markdown_table.append("| " + " | ".join(["---"] * len(header)) + " |")

        # Data rows
        for row in normalized_data[1:]:
            row_clean = [str(cell) if cell is not None else "" for cell in row]
            markdown_table.append("| " + " | ".join(row_clean) + " |")

        return "\n".join(markdown_table)
